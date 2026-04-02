#!/usr/bin/env python3
"""
🔧 DOCTOR LEADS EXCEL HEADER FIXER + CLEANER
Adds exact headers | Removes .0 from phones | NaN → Blank | Professional format
"""

import pandas as pd
import numpy as np
import re
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import argparse

HEADERS = [
    'Search Keyword', 'Category', 'Website Source', 'City', 'District',
    'State', 'Country', 'Phone Number', 'Email Address', 'Doctor Profile Link',
    'Google Maps Link', 'Source URL', 'LinkedIn URL', 'Timestamp'
]

def clean_phone(phone):
    """Remove .0 and format"""
    if pd.isna(phone):
        return ''
    phone_str = str(phone).strip()
    # Remove float artifacts
    phone_str = phone_str.replace('.0', '')
    phone_str = re.sub(r'[^\d\s+()-]', '', phone_str)  # Keep common formats
    return phone_str.strip()

def fix_excel_headers(input_path, output_path):
    print(f"📂 Loading {input_path}")
    
    # Load without headers (use row 0 as data)
    try:
        df = pd.read_excel(input_path, header=None)
        print(f"📊 Loaded {len(df)} rows x {len(df.columns)} columns (no headers)")
    except:
        print("❌ Cannot read Excel - try CSV or check format")
        return
    
    # Assume data starts immediately
    if len(df) == 0:
        print("❌ No data found")
        return
    
    # Create new DF with exact headers
    data = df.values.tolist()
    
    # Clean each column
    cleaned_data = []
    for row in data:
        cleaned_row = []
        for i, cell in enumerate(row):
            if pd.isna(cell) or str(cell).lower() in ['nan', 'none', 'null']:
                cleaned_row.append('')
            elif i == 7:  # Phone column (index 7)
                cleaned_row.append(clean_phone(cell))
            else:
                cleaned_row.append(str(cell).strip())
        cleaned_data.append(cleaned_row)
    
    # Truncate/pad to exact columns
    final_data = []
    for row in cleaned_data:
        row = row[:len(HEADERS)]  # Truncate extra columns
        row += [''] * (len(HEADERS) - len(row))  # Pad missing
        final_data.append(row)
    
    # Create DataFrame
    df_final = pd.DataFrame(final_data, columns=HEADERS)
    
    # Save professionally formatted Excel
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_final.to_excel(writer, sheet_name='Doctor Leads', index=False)
        
        # Format
        worksheet = writer.sheets['Doctor Leads']
        
        # Headers: Bold, blue background
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        
        for col in range(1, len(HEADERS) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            worksheet.row_dimensions[1].height = 25  # Row height
        
        # Auto column width
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze headers
        worksheet.freeze_panes = 'A2'
    
    print(f"✅ SAVED PROFESSIONAL EXCEL: {output_path}")
    print(f"📊 Shape: {len(df_final)} rows x {len(df_final.columns)} columns")
    print("\n📋 HEADERS ADDED:")
    print(" | ".join(HEADERS))
    
    # Preview first row
    if len(df_final) > 0:
        print("\n📋 FIRST ROW:")
        print(" | ".join([str(x)[:30] for x in df_final.iloc[0].tolist()]))
    
    print(f"\n🚀 Ready to upload:")
    print(f"python fixed_google_sheet_uploader_v2.py --file {output_path}")

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="Fix Doctor Excel Headers")
    parser.add_argument('--input', '-i', default='data/doctor_leads_master.xlsx')
    parser.add_argument('--output', '-o', default='data/doctor_leads_fixed.xlsx')
    args = parser.parse_args()
    
    fix_excel_headers(args.input, args.output)


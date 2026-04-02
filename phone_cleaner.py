#!/usr/bin/env python3
"""
DOCTOR LEADS PHONE CLEANER - Multi-Country Intelligent Validation
Input: doctor_leads_master.xlsx -> Output: doctor_leads_verified.xlsx
Rules: Country-based length validation + auto country code addition
"""

import pandas as pd
import re
import numpy as np
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Country validation rules: {country: {'lengths': [valid_lengths], 'code': '+XX'}}
COUNTRY_RULES = {
    'India': {'lengths': [10], 'code': '+91'},
    'USA': {'lengths': [10], 'code': '+1'},
    'United States': {'lengths': [10], 'code': '+1'},
    'UK': {'lengths': [10, 11], 'code': '+44'},
    'United Kingdom': {'lengths': [10, 11], 'code': '+44'},
    'Australia': {'lengths': [9], 'code': '+61'},
    'Canada': {'lengths': [10], 'code': '+1'}
}

def normalize_country(country):
    """Normalize country names for matching"""
    if pd.isna(country):
        return 'India'  # Default fallback
    country_str = str(country).strip().title()
    # Find closest match
    for key in COUNTRY_RULES:
        if key in country_str or country_str in key:
            return key
    return 'India'  # Default

def clean_phone(phone, country):
    """Intelligent phone cleaning with country validation"""
    if pd.isna(phone):
        return ''
    
    # Step 1: Convert to string, remove .0 artifacts
    phone_str = str(phone).strip().rstrip('.0')
    
    # Step 2: Replace nan
    if phone_str.lower() == 'nan':
        return ''
    
    # Step 3: Extract digits (preserve leading + if present)
    digits_only = re.sub(r'[^\d+]', '', phone_str)
    
    if not digits_only or len(digits_only) < 7:
        return ''
    
    # Step 4: Check if already has country code
    norm_country = normalize_country(country)
    if norm_country not in COUNTRY_RULES:
        return ''
    country_rule = COUNTRY_RULES[norm_country]
    country_code = country_rule['code']
    if digits_only.startswith(country_code):
        # Already has country code - validate length (including code)
        full_len = len(digits_only)
        expected_lens = [len(country_code) + L for L in country_rule['lengths']]
        if full_len in expected_lens:
            return digits_only  # Valid, keep unchanged
        else:
            return ''  # Invalid length even with code
    
    # Step 5: No country code - validate raw length and add code
    raw_len = len(digits_only)
    valid_lengths = country_rule['lengths']
    
    if raw_len in valid_lengths:
        return country_code + digits_only  # Valid length, add code
    elif raw_len > max(valid_lengths) + 2:  # Allow minor prefix garbage
        cleaned = digits_only[-max(valid_lengths):]
        if len(cleaned) in valid_lengths:
            return country_code + cleaned
    return ''  # Invalid

def main():
    input_file = 'data/doctor_leads_master.xlsx'
    output_file = 'doctor_leads_verified.xlsx'
    
    input_path = Path(input_file)
    if not input_path.exists():
        print(f'❌ {input_file} not found!')
        print('Available: ls data/*.xlsx *.xlsx')
        return
    
    print(f'📂 Loading: {input_file}')
    df = pd.read_excel(input_file)
    
    print(f'📊 Dataset: {len(df)} rows x {len(df.columns)} cols')
    print(f'📱 Phone before: {df["Phone Number"].notna().sum()}')
    
    if 'Country' not in df.columns:
        print('❌ No Country column!')
        return
    
    # Clean Phone Number column using Country
    df['Phone Number'] = df.apply(
        lambda row: clean_phone(row['Phone Number'], row['Country']), axis=1
    )
    
    # Ensure string dtype (Excel safety)
    df['Phone Number'] = df['Phone Number'].astype(str)
    
    print(f'📱 Phone after: {df["Phone Number"].str.strip().ne("").sum()} valid')
    
    # Stats by country
    stats = df.groupby('Country')['Phone Number'].apply(
        lambda x: x.str.strip().ne('').sum()
    ).to_dict()
    print('📈 Valid phones by country:', stats)
    
    # Save with formatting (preserve all columns unchanged except Phone)
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Doctor Leads', index=False)
        
        # Format headers
        ws = writer.sheets['Doctor Leads']
        for col in ws.iter_rows(min_row=1, max_row=1):
            for cell in col:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill('solid', fgColor='4472C4')
                cell.alignment = Alignment(horizontal='center')
        
        ws.freeze_panes = 'A2'
        
        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            col_letter = column[0].column_letter
            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell.value or '')))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 30)
    
    print(f'✅ SAVED: {output_file}')
    print('🎉 Phone numbers cleaned with country validation!')
    print('Run: open doctor_leads_verified.xlsx')

if __name__ == '__main__':
    main()


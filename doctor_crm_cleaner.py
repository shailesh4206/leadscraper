#!/usr/bin/env python3
"""
🚀 DOCTOR LEADS CRM CLEANER - Production Ready
Fixes: phone .0 → clean | NaN → blank | duplicates | headers | Excel-safe
Input: raw Excel → Output: doctor_leads_crm_ready.xlsx
"""

import pandas as pd
import numpy as np
import re
from pathlib import Path
import argparse
import warnings
warnings.filterwarnings('ignore')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def clean_phone(phone):
    """Convert 9818093267.0 → 9818093267 (string, no scientific notation)"""
    if pd.isna(phone):
        return ''
    # Convert to string, remove .0/float artifacts
    phone_str = str(phone).strip()
    # Remove decimal/float artifacts
    phone_str = re.sub(r'\.0+$', '', phone_str)
    # Clean to digits only (keep +91 prefix if present)
    phone_str = re.sub(r'[^\d+]', '', phone_str)
    # India 10-digit validation
    if len(phone_str) == 10 or len(phone_str) == 12:
        return phone_str
    return phone_str[:10] if len(phone_str) > 10 else phone_str

def clean_dataset(input_file, output_file):
    print(f"📂 Loading raw data: {input_file}")
    
    # Auto-detect headers or assume first row is data
    try:
        # Try with headers first
        df = pd.read_excel(input_file)
        print(f"📊 Loaded {len(df)} rows x {len(df.columns)} columns (with headers)")
    except:
        # Fallback: no headers
        df = pd.read_excel(input_file, header=None)
        print(f"📊 Loaded {len(df)} rows x {len(df.columns)} columns (no headers)")
    
    if len(df) == 0:
        print("❌ No data found")
        return
    
    # Generic headers if none exist
    if df.columns.str.contains('Unnamed').any():
        df.columns = [f'Col_{i}' for i in range(len(df.columns))]
    
    print("🔧 Cleaning...")
    
    # 1. CLEAN PHONES (remove .0, make string)
    phone_cols = df.select_dtypes(include=[np.number]).columns.tolist()
    for col in phone_cols:
        df[col] = df[col].apply(clean_phone)
        print(f"✅ Cleaned phone column: {col}")
    
    # 2. NaN → BLANK
    df = df.fillna('')
    
# 3. DUPLICATE REMOVAL (phone + email columns if exist)
    before_dup = len(df)
    phone_cols = [col for col in ['Phone Number', 'Email', 'Phone', 'Email Address'] if col in df.columns]
    email_cols = [col for col in ['Email', 'Email Address'] if col in df.columns]
    
    dup_cols = phone_cols + email_cols
    if dup_cols:
        df = df.drop_duplicates(subset=dup_cols, keep='first')
        print(f"🗑️ Removed {before_dup - len(df)} duplicates")
    else:
        print("ℹ️ No phone/email columns for dedup")
    
    # 4. EXCEL-SAFE TYPES (no scientific notation)
    for col in df.columns:
        df[col] = df[col].astype(str).str.strip()
    
    # 5. STANDARD CRM COLUMNS (pad/trim)
    crm_columns = [
        'Doctor Name', 'Specialty', 'Clinic/Hospital', 'City', 'Phone Number', 
        'Email', 'Website', 'Source', 'Timestamp'
    ]
    
    # Map existing columns to CRM (flexible)
    df_crm = pd.DataFrame()
    for crm_col in crm_columns:
        if crm_col.lower() in df.columns.str.lower():
            df_crm[crm_col] = df[df.columns[df.columns.str.lower().str.contains(crm_col.lower())]].iloc[:, 0]
        else:
            df_crm[crm_col] = ''
    
    # Pad extra columns if needed
    extra_cols = [col for col in df.columns if col not in df_crm.columns]
    for col in extra_cols[:3]:  # First 3 extra
        df_crm[f'Extra_{col}'] = df[col]
    
# Remove empty rows (phone OR email)
    mask = (df_crm['Phone Number'].str.strip() != '') | (df_crm['Email'].str.strip() != '')
    df_crm = df_crm[mask]
    
    print(f"📊 Final dataset: {len(df_crm)} CRM-ready rows")
    
    # SAVE EXCEL-SAFE
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_crm.to_excel(writer, sheet_name='CRM Leads', index=False)
        
        # Professional formatting
        ws = writer.sheets['CRM Leads']
        
        # Headers
        for col in ws.iter_rows(min_row=1, max_row=1):
            for cell in col:
                cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
                cell.fill = openpyxl.styles.PatternFill("solid", fgColor="4472C4")
                cell.alignment = openpyxl.styles.Alignment(horizontal="center")
        
        # Freeze headers
        ws.freeze_panes = 'A2'
        
        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 25)
    
    print(f"✅ SAVED: {output_file}")
    
    # PREVIEW
    print("\n📋 CRM-READY PREVIEW:")
    print(df_crm[['Doctor Name', 'Phone Number', 'Email', 'City']].head().to_string())
    
    # STATS
    print("\n📈 STATS:")
    print(f"• Rows: {len(df_crm)}")
    print(f"• Phones: {df_crm['Phone Number'].str.strip().ne('').sum()}")
    print(f"• Emails: {df_crm['Email'].str.strip().ne('').sum()}")
    print(f"• Cities: {df_crm['City'].nunique()}")

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--input', default='data/doctor_leads_master.xlsx')
    parser.add_argument('--output', default='data/doctor_leads_crm_ready.xlsx')
    args = parser.parse_args()
    
    clean_dataset(args.input, args.output)
    
    print("\n🚀 READY FOR CRM / Google Sheets!")
    print(f"Upload: python google_sheet_uploader.py --file {args.output}")

